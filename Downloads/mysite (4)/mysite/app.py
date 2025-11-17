# ==============================================================================
# Imports
# ==============================================================================
from flask import Flask, request, abort, render_template, jsonify, redirect, url_for, send_file
from google.cloud.firestore_v1.base_query import FieldFilter, Or
from linebot import LineBotApi, WebhookHandler
from linebot.exceptions import InvalidSignatureError
from linebot.models import MessageEvent, TextMessage, TextSendMessage, FlexSendMessage, BubbleContainer, BoxComponent, ButtonComponent, URIAction, TextComponent, ImageMessage, VideoMessage, AudioMessage, FileMessage
from datetime import datetime, timedelta
import os
import logging
import firebase_admin
from firebase_admin import credentials, firestore, storage
import requests
import sys
import uuid
import tempfile
from functools import wraps
import openai
import random
import re
import string
from dotenv import load_dotenv
import google.generativeai as genai
import json
# Load environment variables from .env file
load_dotenv()

# ==============================================================================
# Configuration and Initialization
# ==============================================================================
# OpenAI API Key
openai.api_key = os.getenv("OPENAI_API_KEY")

# Gemini API Key
GEMINI_API_KEY = os.getenv("GEMINI_API_KEY")
if GEMINI_API_KEY:
    genai.configure(api_key=GEMINI_API_KEY)

# NGワードリスト (簡易的な実装)
NG_WORDS = ["死ね", "殺す", "バカ", "アホ", "消えろ"] # 必要に応じて追加・変更

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# ログ設定
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# LINE Bot 設定
LINE_CHANNEL_SECRET = "fe96005b5be582c09f6d8c192fbf6b06"
LINE_CHANNEL_ACCESS_TOKEN = 'SvfdWFWa+A+nrZTLlBGMPKoEf6fN/miJg93DS0benY9Jb/6fjh/5PBD/Wrz/RKJ2TLGoHhDUuGOee4hQJsHQ5gwZP0elG25qUPp8pzVbZCvhsoY7UHjki20EeU/fN7xhy07hUuJQIpdQtojgbp/pPAdB04t89/1O/w1cDnyilFU='

# LINE Login 設定 (LIFF認証用)
LINE_LOGIN_CHANNEL_ID = "2008454581"
LINE_LOGIN_CHANNEL_SECRET = "f4999a4e5ed14c42c92871ffc5a01d39"

# LIFF ID (単一のLIFFアプリを使用)
LIFF_ID_PRIMARY = "2008454581-9AVyN4Jv"

# FlaskとLINE SDKの初期化
app = Flask(__name__)
line_bot_api = LineBotApi(LINE_CHANNEL_ACCESS_TOKEN)
handler = WebhookHandler(LINE_CHANNEL_SECRET)

# Firestore設定 (サービスアカウント認証)
FIREBASE_KEY_FILENAME = 'firebase-key.json'
FIREBASE_KEY_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), FIREBASE_KEY_FILENAME)

try:
    # Firebaseの初期化
    if not firebase_admin._apps:
        cred = credentials.Certificate(FIREBASE_KEY_PATH)
        firebase_admin.initialize_app(cred, {
            'projectId': 'satounikikun',
            'storageBucket': 'satounikikun.firebasestorage.app'
        })

    db = firestore.client()
    bucket = storage.bucket('satounikikun.firebasestorage.app')
    logger.info("Firebase and Firestore connection successful.")
except Exception as e:
    logger.error(f"Firestore initialization failed: {e}")
    db = None
    bucket = None

# ==============================================================================
# Decorators
# ==============================================================================
def token_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        id_token = None
        if 'Authorization' in request.headers and request.headers['Authorization'].startswith('Bearer '):
            id_token = request.headers['Authorization'].split(' ')[1]
        elif 'idToken' in request.form:
            id_token = request.form.get('idToken')
        elif request.is_json and 'idToken' in request.get_json():
            id_token = request.get_json().get('idToken')

        if not id_token:
            return jsonify({"status": "error", "message": "ID Token is missing"}), 400

        try:
            logger.info("Attempting token verification.") # Fixed log message
            logger.info(f"Server current UTC time before token verification: {datetime.utcnow().isoformat()}Z")
            res = requests.post('https://api.line.me/oauth2/v2.1/verify', data={
                'id_token': id_token,
                'client_id': LINE_LOGIN_CHANNEL_ID
            }) # PROXY REMOVED

            if res.status_code != 200:
                error_response = res.json()
                error_description = error_response.get('error_description', 'Unknown verification error.')
                logger.error(f"ID Token verification failed with status {res.status_code}: {res.text}")

                # クライアントに返すエラーメッセージを具体的にする
                message = "ID Token verification failed."
                if 'expired' in error_description.lower():
                    message = "ID Token has expired. Please re-authenticate."

                return jsonify({
                    "status": "error",
                    "message": message,
                    "detail": error_description
                }), 401

            token_info = res.json()
            line_user_id = token_info.get('sub')

            if not line_user_id:
                logger.error("Verified ID Token does not contain 'sub' (user ID).")
                return jsonify({"status": "error", "message": "Invalid ID Token (no user ID)"}), 401

            return f(line_user_id, *args, **kwargs)

        except requests.exceptions.RequestException as e:
            logger.error(f"Request to LINE verify endpoint failed: {e}")
            return jsonify({"status": "error", "message": "ID Token verification request failed"}), 500
        except Exception as e:
            logger.error(f"Error processing ID Token: {e}")
            return jsonify({"status": "error", "message": "Internal server error during token processing"}), 500

    return decorated_function

# ==============================================================================
# Helper Functions
# ==============================================================================
def create_user_if_not_exists(user_id):
    """
    指定されたuser_idのユーザーが存在しない場合、LINEプロファイルから情報を取得してFirestoreに作成します。
    """
    try:
        user_ref = db.collection('users').where(filter=FieldFilter('line_user_id', '==', user_id)).limit(1)
        docs = list(user_ref.stream())

        if not docs:
            profile = line_bot_api.get_profile(user_id)
            display_name = profile.display_name
            new_user_data = {
                'line_user_id': user_id,
                'name': display_name,
                'school': '',
                'class_name': '',
                'icon_path': '',
                'is_registered': False,
                'role': 'student',
                'class_token_id': '',
                'is_posting_diary': False,
                'created_at': datetime.now().isoformat()
            }
            db.collection('users').add(new_user_data)
            logger.info(f"New user created: {display_name} (ID: {user_id}) with role 'student'")

    except Exception as e:
        logger.error(f"Failed to create or check user: {e}")

# ==============================================================================
# LINE Webhook
# ==============================================================================
@app.route("/callback", methods=['POST'])
def callback():
    signature = request.headers.get('X-Line-Signature')
    body = request.get_data(as_text=True)

    try:
        handler.handle(body, signature)
    except InvalidSignatureError:
        abort(400)
    except Exception as e:
        abort(500)

    return 'OK', 200

@handler.add(MessageEvent, message=TextMessage)
def handle_message(event):
    user_id = event.source.user_id
    user_message = event.message.text
    logger.info(f"Received message from user {user_id}: '{user_message}'")

    user_message_upper = user_message.upper()
    class_code_to_join = None

    # "join" または "参加" で始まるか、6桁の英数字コードそのものであるかをチェック
    logger.info(f"Checking message '{user_message_upper}' for join commands.")
    is_join = user_message_upper.startswith("JOIN ")
    is_sanka = user_message_upper.startswith("参加 ")
    is_code_only = bool(re.fullmatch(r'[A-Z0-9]{6}', user_message_upper))
    logger.info(f"startswith('JOIN '): {is_join}, startswith('参加 '): {is_sanka}, fullmatch(code): {is_code_only}")

    if is_join:
        class_code_to_join = user_message_upper[5:].strip()
        logger.info(f"Join command detected. Extracted code: '{class_code_to_join}'")
    elif is_sanka:
        class_code_to_join = user_message_upper[3:].strip()
        logger.info(f"Sanka command detected. Extracted code: '{class_code_to_join}'")
    elif is_code_only:
        class_code_to_join = user_message_upper
        logger.info(f"Code-only message detected: '{class_code_to_join}'")

    if class_code_to_join:
        logger.info(f"Attempting to join class with code: '{class_code_to_join}' from LINE message.")
        try:
            classes_ref = db.collection('classes').where(filter=FieldFilter('class_code', '==', class_code_to_join)).limit(1)
            class_docs = list(classes_ref.stream())
            if not class_docs:
                reply_text = f"無効なクラスコードです: {class_code_to_join}"
                line_bot_api.reply_message(event.reply_token, TextSendMessage(text=reply_text))
                return

            class_doc = class_docs[0]
            class_data = class_doc.to_dict()
            class_id = class_doc.id
            class_name = class_data.get('class_name')

            users_ref = db.collection('users')
            user_query = users_ref.where(filter=FieldFilter('line_user_id', '==', user_id)).limit(1)
            user_docs_list = list(user_query.stream())

            user_doc = None
            user_data = {}
            if user_docs_list:
                user_doc = user_docs_list[0]
                user_data = user_doc.to_dict()

            # 既に参加済み、または申請中かチェック
            if 'class_memberships' in user_data:
                for membership in user_data['class_memberships']:
                    if membership.get('class_id') == class_id:
                        if membership.get('status') == 'approved':
                            reply_text = f"既に「{class_name}」に参加しています。"
                            line_bot_api.reply_message(event.reply_token, TextSendMessage(text=reply_text))
                            return
                        elif membership.get('status') == 'pending':
                            reply_text = "このクラスには既に申請済みです。先生の承認をお待ちください。"
                            line_bot_api.reply_message(event.reply_token, TextSendMessage(text=reply_text))
                            return

            new_membership = {
                'class_id': class_id,
                'class_name': class_name,
                'status': 'pending',
                'requested_at': datetime.now().isoformat()
            }

            if user_doc:
                user_doc.reference.update({
                    'class_memberships': firestore.ArrayUnion([new_membership]),
                    'pending_class_ids': firestore.ArrayUnion([class_id])
                })
            else:
                profile = line_bot_api.get_profile(user_id)
                display_name = profile.display_name
                new_user_data = {
                    'line_user_id': user_id, 'name': display_name, 'role': 'student',
                    'is_registered': True, 'created_at': datetime.now().isoformat(),
                    'class_memberships': [new_membership], 'pending_class_ids': [class_id]
                }
                users_ref.add(new_user_data)

            reply_text = f"クラス「{class_name}」への参加を申請しました。先生の承認をお待ちください。"
            line_bot_api.reply_message(event.reply_token, TextSendMessage(text=reply_text))
            return

        except Exception as e:
            logger.error(f"join_class from LINE message error for user {user_id} with code {class_code_to_join}: {e}", exc_info=True)
            line_bot_api.reply_message(event.reply_token, TextSendMessage(text="クラス参加中にエラーが発生しました。"))
            return
    else:
        logger.info("Message is not a join command. Proceeding to other handlers.")

    if not db:
        line_bot_api.reply_message(event.reply_token, TextSendMessage(text="エラー：サーバーがデータベースに接続できませんでした。"))
        return

    create_user_if_not_exists(user_id)

    user_doc_ref = db.collection('users').where(filter=FieldFilter('line_user_id', '==', user_id)).limit(1)
    user_docs = list(user_doc_ref.get())
    user_data = {}
    if user_docs:
        user_data = user_docs[0].to_dict()

    if not user_data.get('is_registered', False):
        reply_text = "アカウントを作成するには、先生から配布されるQRコードをスキャンしてクラスに参加してください。"
        line_bot_api.reply_message(event.reply_token, TextSendMessage(text=reply_text))
        return

    if user_message == "クラスのページ":
        reply_text = f"""あなたのクラスページはこちらです。
line://app/{LIFF_ID_PRIMARY}/class_home"""
        line_bot_api.reply_message(event.reply_token, TextSendMessage(text=reply_text))
        return

    elif user_message == "マイページ":
        reply_text = f"""あなたのマイページはこちらです。
line://app/{LIFF_ID_PRIMARY}/mypage"""
        line_bot_api.reply_message(event.reply_token, TextSendMessage(text=reply_text))
        return

    elif user_message == "先生ダッシュボード" and user_data.get('role') == 'teacher':
        reply_text = f"""先生用ダッシュボードはこちらです。
line://app/{LIFF_ID_PRIMARY}/teacher_dashboard"""
        line_bot_api.reply_message(event.reply_token, TextSendMessage(text=reply_text))
        return

    elif user_message == "課題一覧":
        try:
            # Get user's approved classes
            approved_class_ids = [
                m['class_id'] for m in user_data.get('class_memberships', []) if m.get('status') == 'approved'
            ]

            if not approved_class_ids:
                line_bot_api.reply_message(event.reply_token, TextSendMessage(text="参加中のクラスがありません。"))
                return

            # For simplicity, let's just use the first approved class for now.
            # A more advanced implementation would let the user choose or show all.
            target_class_id = approved_class_ids[0]

            # Fetch assignments for the class
            assignments_ref = db.collection('assignments').where(
                filter=FieldFilter('class_id', '==', target_class_id)
            ).order_by('due_date', direction=firestore.Query.ASCENDING)
            assignments_docs = list(assignments_ref.stream())

            if not assignments_docs:
                line_bot_api.reply_message(event.reply_token, TextSendMessage(text="現在、提出する課題はありません。"))
                return

            # Fetch user's submissions to filter out completed assignments
            submissions_ref = db.collection('submissions').where(
                filter=FieldFilter('student_line_user_id', '==', user_id)
            )
            submissions_docs = list(submissions_ref.stream())
            submitted_assignment_ids = {doc.to_dict().get('assignment_id') for doc in submissions_docs}

            pending_assignments = [
                doc.to_dict() for doc in assignments_docs if doc.id not in submitted_assignment_ids
            ]
            
            now = datetime.now().isoformat()
            pending_assignments = [a for a in pending_assignments if a.get('due_date', '') > now]


            if not pending_assignments:
                line_bot_api.reply_message(event.reply_token, TextSendMessage(text="提出期限内の未提出課題はありません。"))
                return

            reply_text = "未提出の課題一覧です。\n提出するには「課題提出 [課題ID]」と送ってください。\n\n"
            for assign in pending_assignments:
                due_date = datetime.fromisoformat(assign['due_date']).strftime('%Y年%m月%d日 %H:%M')
                reply_text += f"■ {assign['title']}\n"
                reply_text += f"ID: {assign['id']}\n"
                reply_text += f"期限: {due_date}\n\n"

            line_bot_api.reply_message(event.reply_token, TextSendMessage(text=reply_text.strip()))

        except Exception as e:
            logger.error(f"Error fetching assignments for user {user_id}: {e}", exc_info=True)
            line_bot_api.reply_message(event.reply_token, TextSendMessage(text="課題一覧の取得中にエラーが発生しました。"))
        return

    elif user_message == "その他":
        reply_text = "どの項目を見ますか？"
        flex_menu = FlexSendMessage(
            alt_text="メニュー",
            contents=BubbleContainer(
                body=BoxComponent(
                    layout="vertical",
                    contents=[
                        BoxComponent(
                            layout="vertical",
                            contents=[
                                BoxComponent(
                                    layout="horizontal",
                                    spacing="md",
                                    contents=[
                                        ButtonComponent(
                                            style="primary",
                                            flex=1,
                                            action=URIAction(
                                                label="提出物",
                                                uri=f"line://app/{LIFF_ID_PRIMARY}/homework"
                                            )
                                        ),
                                        ButtonComponent(
                                            style="primary",
                                            flex=1,
                                            action=URIAction(
                                                label="自己理解・評価",
                                                uri=f"line://app/{LIFF_ID_PRIMARY}/score"
                                            )
                                        )
                                    ]
                                ),
                                BoxComponent(
                                    layout="horizontal",
                                    spacing="md",
                                    margin="md",
                                    contents=[
                                        ButtonComponent(
                                            style="secondary",
                                            flex=1,
                                            action=URIAction(
                                                label="規約・ルール",
                                                uri=f"line://app/{LIFF_ID_PRIMARY}/rules"
                                            )
                                        ),
                                        ButtonComponent(
                                            style="secondary",
                                            flex=1,
                                            action=URIAction(
                                                label="お問い合わせ",
                                                uri=f"line://app/{LIFF_ID_PRIMARY}/contact"
                                            )
                                        )
                                    ]
                                )
                            ]
                        )
                    ]
                )
            )
        )
        line_bot_api.reply_message(event.reply_token, flex_menu)
        return

    else:
        # Check user state for ongoing actions
        user_state = user_data.get('user_state')
        if user_state and user_state.get('action') == 'submitting_assignment':
            assignment_id = user_state.get('assignment_id')
            assignment_title = user_state.get('assignment_title')
            
            try:
                # Save the submission (text only for now)
                submission_ref = db.collection('submissions').document()
                submission_data = {
                    'id': submission_ref.id,
                    'assignment_id': assignment_id,
                    'student_line_user_id': user_id,
                    'submission_type': 'text',
                    'content': user_message,
                    'submitted_at': datetime.now().isoformat()
                }
                submission_ref.set(submission_data)

                # Clear user state
                user_ref = db.collection('users').where(filter=FieldFilter('line_user_id', '==', user_id)).limit(1).get()[0].reference
                user_ref.update({'user_state': firestore.DELETE_FIELD})

                reply_text = f"課題「{assignment_title}」を提出しました。"
                line_bot_api.reply_message(event.reply_token, TextSendMessage(text=reply_text))

            except Exception as e:
                logger.error(f"Error saving submission for user {user_id}: {e}", exc_info=True)
                line_bot_api.reply_message(event.reply_token, TextSendMessage(text="提出物の保存中にエラーが発生しました。"))
            return

        # Check for submission start command
        submission_match = re.match(r'課題提出\s+([a-zA-Z0-9\-_]+)', user_message)
        if submission_match:
            assignment_id = submission_match.group(1)
            try:
                # Verify assignment exists and is not past due
                assignment_ref = db.collection('assignments').document(assignment_id)
                assignment_doc = assignment_ref.get()
                if not assignment_doc.exists:
                    line_bot_api.reply_message(event.reply_token, TextSendMessage(text="指定された課題IDが見つかりません。"))
                    return

                assignment_data = assignment_doc.to_dict()
                now = datetime.now().isoformat()
                if assignment_data.get('due_date', '') < now:
                    line_bot_api.reply_message(event.reply_token, TextSendMessage(text="この課題は提出期限を過ぎています。"))
                    return

                # Set user state
                user_ref = db.collection('users').where(filter=FieldFilter('line_user_id', '==', user_id)).limit(1).get()[0].reference
                user_ref.update({
                    'user_state': {
                        'action': 'submitting_assignment',
                        'assignment_id': assignment_id,
                        'assignment_title': assignment_data.get('title')
                    }
                })

                reply_text = f"課題「{assignment_data.get('title')}」の提出内容を送信してください。テキストまたはファイルを送信できます。"
                line_bot_api.reply_message(event.reply_token, TextSendMessage(text=reply_text))

            except Exception as e:
                logger.error(f"Error starting assignment submission for user {user_id}: {e}", exc_info=True)
                line_bot_api.reply_message(event.reply_token, TextSendMessage(text="課題提出の準備中にエラーが発生しました。"))
            return

        # 不適切ワードチェック
        for ng_word in NG_WORDS:
            if ng_word in user_message:
                reply_text = f"""不適切な言葉が含まれています。日記は保存されませんでした。
「{ng_word}」のような言葉は使用しないでください。"""
                line_bot_api.reply_message(event.reply_token, TextSendMessage(text=reply_text))
                return
        
        # ユーザードキュメント取得
        user_docs = db.collection('users').where(filter=FieldFilter('line_user_id', '==', user_id)).limit(1).get()
        if not user_docs:
            reply_text = "ユーザー情報が見つかりません。"
            line_bot_api.reply_message(event.reply_token, TextSendMessage(text=reply_text))
            return

        user_ref = user_docs[0].reference
        user_data = user_docs[0].to_dict()
        
        # 🟢 「日記を投稿します」モード開始
        if user_message == "日記を投稿します":
            user_ref.update({'is_posting_diary': True})
            user_data['is_posting_diary'] = True  # ←これを追加！
            reply_text = "📝 次に送るメッセージを日記として保存します。"
            line_bot_api.reply_message(event.reply_token, TextSendMessage(text=reply_text))
            return

        # 🟢 投稿モード中なら、次のメッセージを日記として保存
        if user_data.get('is_posting_diary', False):
            try:
                approved_memberships = [m for m in user_data.get('class_memberships', []) if m.get('status') == 'approved']

                target_class_id = None
                if len(approved_memberships) == 1:
                    target_class_id = approved_memberships[0].get('class_id')
                elif len(approved_memberships) > 1:
                    # 先生の場合、直近で作成したクラスをデフォルトとして使用する
                    if user_data.get('role') == 'teacher':
                        classes_ref = db.collection('classes') \
                            .where(filter=FieldFilter('teacher_line_user_id', '==', user_id)) \
                            .order_by('created_at', direction=firestore.Query.DESCENDING) \
                            .limit(1)
                        class_docs = list(classes_ref.stream())
                        if class_docs:
                            target_class_id = class_docs[0].id
                    else:
                        reply_text = "複数のクラスに参加しています。日記を投稿するクラスをWebアプリのクラスホーム画面で選択してから投稿してください。"
                        line_bot_api.reply_message(event.reply_token, TextSendMessage(text=reply_text))
                        return

                if not target_class_id:
                    reply_text = "参加が承認されたクラスがありません。日記を投稿できません。"
                    line_bot_api.reply_message(event.reply_token, TextSendMessage(text=reply_text))
                    return

                # Firestoreに日記を保存
                diary_data = {
                    'user_id': user_id,
                    'content': user_message,
                    'class_id': target_class_id, # class_join_tokenからclass_idに変更
                    'created_at': datetime.now().isoformat()
                }
                db.collection('diaries').add(diary_data)

                # 投稿モードを終了
                user_ref.update({'is_posting_diary': False})

                reply_text = """✅ 日記を保存しました！
また投稿するときは「日記を投稿します」と送ってください。"""
                logger.info(f"Diary saved for user {user_id} in class {target_class_id}.")
            except Exception as e:
                logger.error(f"Failed to save diary for user {user_id}: {e}")
                reply_text = "日記の保存に失敗しました。もう一度お試しください。"

            line_bot_api.reply_message(event.reply_token, TextSendMessage(text=reply_text))
            return

        # ⚪️ それ以外の通常メッセージ
        else:
            reply_text = """📘 コマンドが認識されませんでした。
「日記を投稿します」と送ってみてください。"""
            line_bot_api.reply_message(event.reply_token, TextSendMessage(text=reply_text))
            return

@handler.add(MessageEvent, message=[ImageMessage, VideoMessage, AudioMessage, FileMessage])
def handle_content_message(event):
    user_id = event.source.user_id

    # Get user data and check state
    user_docs = db.collection('users').where(filter=FieldFilter('line_user_id', '==', user_id)).limit(1).get()
    if not user_docs:
        # Or handle as an error
        return

    user_ref = user_docs[0].reference
    user_data = user_docs[0].to_dict()
    user_state = user_data.get('user_state')

    if user_state and user_state.get('action') == 'submitting_assignment':
        assignment_id = user_state.get('assignment_id')
        assignment_title = user_state.get('assignment_title')

        try:
            message_content = line_bot_api.get_message_content(event.message.id)
            
            with tempfile.NamedTemporaryFile(delete=False) as tf:
                for chunk in message_content.iter_content():
                    tf.write(chunk)
                temp_path = tf.name

            # Upload to Firebase Storage
            file_extension = ''
            filename = ''
            if isinstance(event.message, FileMessage):
                filename = event.message.file_name
                file_extension = os.path.splitext(filename)[1]

            unique_filename = f"submissions/{assignment_id}/{user_id}_{datetime.now().strftime('%Y%m%d%H%M%S')}{file_extension}"
            blob = bucket.blob(unique_filename)
            
            content_type = None
            if isinstance(event.message, ImageMessage):
                content_type = 'image/jpeg' # Assuming jpeg, but could be other types
            elif isinstance(event.message, VideoMessage):
                content_type = 'video/mp4'
            elif isinstance(event.message, AudioMessage):
                content_type = 'audio/mp3'
            elif hasattr(event.message, 'file_name'):
                 # For FileMessage, try to guess from extension
                 pass # No easy way to get content type from sdk

            blob.upload_from_filename(temp_path) # Let Firebase guess content type
            os.unlink(temp_path)

            blob.make_public()
            public_url = blob.public_url

            # Save submission record to Firestore
            submission_ref = db.collection('submissions').document()
            submission_data = {
                'id': submission_ref.id,
                'assignment_id': assignment_id,
                'student_line_user_id': user_id,
                'submission_type': 'file',
                'content': public_url, # URL to the file
                'file_name': filename,
                'submitted_at': datetime.now().isoformat()
            }
            submission_ref.set(submission_data)

            # Clear user state
            user_ref.update({'user_state': firestore.DELETE_FIELD})

            reply_text = f"課題「{assignment_title}」としてファイルを提出しました。"
            line_bot_api.reply_message(event.reply_token, TextSendMessage(text=reply_text))

        except Exception as e:
            logger.error(f"Error saving file submission for user {user_id}: {e}", exc_info=True)
            line_bot_api.reply_message(event.reply_token, TextSendMessage(text="ファイル提出の処理中にエラーが発生しました。"))
        return

# ==============================================================================
# API Endpoints
# ==============================================================================
@app.route('/api/user/upload_icon', methods=['POST'])
@token_required
def upload_icon(uploader_user_id):
    if not db or not bucket:
        print("Firestore or Firebase Storage is not initialized.", file=sys.stderr)
        return jsonify({"status": "error", "message": "Database or Storage connection failed"}), 500

    if 'icon' not in request.files:
        return jsonify({"status": "error", "message": "No icon file provided"}), 400

    icon_file = request.files['icon']
    if icon_file.filename == '':
        return jsonify({"status": "error", "message": "No selected file"}), 400

    try:
        original_filename = icon_file.filename
        file_extension = os.path.splitext(original_filename)[1]
        unique_filename = f"{uploader_user_id}/{datetime.now().strftime('%Y%m%d%H%M%S')}_{uuid.uuid4().hex}{file_extension}"

        blob = bucket.blob(unique_filename)
        blob.upload_from_file(icon_file, content_type=icon_file.content_type)

        blob.make_public()
        public_url = blob.public_url

        return jsonify({"status": "success", "icon_path": public_url}), 200

    except Exception as e:
        print(f"Error uploading icon: {e}", file=sys.stderr)
        return jsonify({"status": "error", "message": "Failed to upload icon"}), 500

@app.route('/api/teacher/classes', methods=['GET'])
@token_required
def get_classes(teacher_line_user_id):
    if not db:
        return jsonify({"status": "error", "message": "Database connection failed"}), 500

    try:
        classes_ref = db.collection('classes').where(filter=FieldFilter('teacher_line_user_id', '==', teacher_line_user_id)).order_by('created_at', direction=firestore.Query.DESCENDING)
        docs = classes_ref.stream()

        class_list = []
        for doc in docs:
            class_data = doc.to_dict()
            join_url = f"line://app/{LIFF_ID_PRIMARY}/join_class?token={class_data.get('join_token')}"

            class_list.append({
                'id': doc.id,
                'class_name': class_data.get('class_name'),
                'class_code': class_data.get('class_code'),
                'join_url': join_url,
                'created_at': class_data.get('created_at')
            })

        return jsonify({"status": "success", "data": class_list}), 200

    except Exception as e:
        logger.error(f"Error fetching classes for teacher {teacher_line_user_id}: {e}")
        return jsonify({"status": "error", "message": "Failed to fetch classes"}), 500

def generate_class_code(length=6):
    """ランダムなクラスコードを生成"""
    letters_and_digits = string.ascii_uppercase + string.digits
    return ''.join(random.choice(letters_and_digits) for i in range(length))

@app.route('/api/teacher/classes', methods=['POST'])
@token_required
def create_class(teacher_line_user_id):
    if not db:
        return jsonify({"status": "error", "message": "Database connection failed"}), 500

    data = request.get_json()
    class_name = data.get('class_name')

    if not class_name:
        return jsonify({"status": "error", "message": "Class name is required"}), 400

    teacher_user_doc = db.collection('users').where(filter=FieldFilter('line_user_id', '==', teacher_line_user_id)).limit(1).get()
    if not teacher_user_doc or teacher_user_doc[0].to_dict().get('role') != 'teacher':
        return jsonify({"status": "error", "message": "Unauthorized: Only teachers can create classes"}), 403

    try:
        join_token = str(uuid.uuid4())
        class_code = generate_class_code()

        new_class_data = {
            'class_name': class_name,
            'class_code': class_code,
            'teacher_line_user_id': teacher_line_user_id,
            'join_token': join_token,
            'created_at': datetime.now().isoformat()
        }
        db.collection('classes').add(new_class_data)

        return jsonify({"status": "success", "message": "Class created successfully", "data": new_class_data}), 201

    except Exception as e:
        logger.error(f"Error creating class for teacher {teacher_line_user_id}: {e}")
        return jsonify({"status": "error", "message": "Failed to create class"}), 500

@app.route('/api/student/join_class', methods=['POST'])
@token_required
def join_class(student_line_user_id):
    if not db:
        logger.error("join_class: Firestore is not initialized.")
        return jsonify({"status": "error", "message": "データベース接続エラー"}), 500

    data = request.get_json()
    if not data:
        logger.warning("join_class: No JSON data received.")
        return jsonify({"status": "error", "message": "リクエストが空です。"}), 400

    class_code_input = data.get('class_code')
    logger.info(f"join_class: Attempting to join with class_code: {class_code_input}")

    if not class_code_input:
        return jsonify({"status": "error", "message": "クラスコードがありません。"}), 400

    try:
        classes_ref = db.collection('classes').where(filter=FieldFilter("class_code", "==", class_code_input)).limit(1)
        class_docs = list(classes_ref.stream())

        if not class_docs:
            return jsonify({"status": "error", "message": "無効なクラスコードです。"}), 404

        class_doc = class_docs[0]
        class_data = class_doc.to_dict()
        class_id = class_doc.id
        class_name = class_data.get('class_name')

        users_ref = db.collection('users')
        user_query = users_ref.where(filter=FieldFilter("line_user_id", "==", student_line_user_id)).limit(1)
        user_docs_list = list(user_query.stream())

        user_doc = None
        user_data = {}
        if user_docs_list:
            user_doc = user_docs_list[0]
            user_data = user_doc.to_dict()

        # 既に参加済み、または申請中かチェック
        if 'class_memberships' in user_data:
            for membership in user_data['class_memberships']:
                if membership.get('class_id') == class_id:
                    if membership.get('status') == 'approved':
                        return jsonify({"status": "error", "message": f"既に「{class_name}」に参加しています。"}), 409
                    elif membership.get('status') == 'pending':
                        return jsonify({"status": "error", "message": "このクラスには既に申請済みです。先生の承認をお待ちください。"}), 409

        new_membership = {
            'class_id': class_id,
            'class_name': class_name,
            'status': 'pending',
            'requested_at': datetime.now().isoformat()
        }

        if user_doc:
            user_doc.reference.update({
                'class_memberships': firestore.ArrayUnion([new_membership]),
                'pending_class_ids': firestore.ArrayUnion([class_id])
            })
            logger.info(f"User {student_line_user_id} requested to join class {class_name}")
        else:
            # 新規ユーザーの場合は、ユーザー情報も一緒に作成
            profile = line_bot_api.get_profile(student_line_user_id)
            display_name = profile.display_name
            new_user_data = {
                'line_user_id': student_line_user_id,
                'name': display_name,
                'role': 'student',
                'is_registered': True,
                'created_at': datetime.now().isoformat(),
                'class_memberships': [new_membership],
                'pending_class_ids': [class_id]
            }
            users_ref.add(new_user_data)
            logger.info(f"New user {student_line_user_id} created and requested to join class {class_name}")

        return jsonify({"status": "success", "message": f"クラス「{class_name}」への参加を申請しました。先生の承認をお待ちください。"}), 200

    except Exception as e:
        logger.error(f"Error in join_class with class_code {class_code_input}: {e}", exc_info=True)
        return jsonify({"status": "error", "message": "サーバーエラーが発生しました。"}), 500

@app.route('/api/user', methods=['POST'])
@token_required
def update_user_profile(line_user_id):
    if not db:
        print("Firestore is not initialized.", file=sys.stderr)
        return jsonify({"status": "error", "message": "Database connection failed"}), 500

    data = request.get_json()
    name = data.get('name')
    school = data.get('school')
    class_name = data.get('class')

    try:
        users_ref = db.collection('users')
        query = users_ref.where(filter=FieldFilter('line_user_id', '==', line_user_id)).limit(1)
        docs = query.stream()

        user_doc_id = None
        for doc in docs:
            user_doc_id = doc.id
            break

        if user_doc_id:
            update_data = {
                'name': name,
                'school': school,
                'class_name': class_name,
                'icon_path': data.get('icon_path', ''),
                'is_registered': True,
                'updated_at': datetime.now().isoformat()
            }
            db.collection('users').document(user_doc_id).update(update_data)
            return jsonify({"status": "success", "message": "Profile updated successfully"}), 200
        else:
            new_user_data = {
                'line_user_id': line_user_id,
                'name': name,
                'school': school,
                'class_name': class_name,
                'icon_path': data.get('icon_path', ''),
                'is_registered': True,
                'role': 'student',
                'is_posting_diary': False,
                'created_at': datetime.now().isoformat(),
                'updated_at': datetime.now().isoformat()
            }
            db.collection('users').add(new_user_data)
            return jsonify({"status": "success", "message": "Profile created successfully"}), 201

    except Exception as e:
        print(f"Error updating user profile in Firestore: {e}", file=sys.stderr)
        return jsonify({"status": "error", "message": "Failed to update profile"}), 500

@app.route('/api/user', methods=['GET'])
@token_required
def get_user_profile(line_user_id):
    if not db:
        print("Firestore is not initialized.", file=sys.stderr)
        return jsonify({"status": "error", "message": "Database connection failed"}), 500

    try:
        users_ref = db.collection('users')
        query = users_ref.where(filter=FieldFilter('line_user_id', '==', line_user_id)).limit(1)
        docs = query.stream()

        user_data = None
        for doc in docs:
            user_data = doc.to_dict()
            break

        if user_data:
            # is_registeredはclass_membershipsに承認済みのものがあるかで判断
            approved_memberships = [m for m in user_data.get('class_memberships', []) if m.get('status') == 'approved']
            is_registered = len(approved_memberships) > 0

            response_data = {
                'name': user_data.get('name', ''),
                'school': user_data.get('school', ''),
                'icon_path': user_data.get('icon_path', ''),
                'is_registered': is_registered,
                'role': user_data.get('role', 'student'),
                'class_memberships': user_data.get('class_memberships', []),
                'settings': user_data.get('settings', {})
            }
            return jsonify({"status": "success", "data": response_data}), 200
        else:
            # LIFFで認証済みだが、DBにユーザーがいない場合（起こりにくいが念のため）
            return jsonify({"status": "error", "message": "User profile not found"}), 404

    except Exception as e:
        print(f"Error fetching user profile from Firestore: {e}", file=sys.stderr)
        return jsonify({"status": "error", "message": "Failed to fetch profile"}), 500

@app.route('/api/user/settings', methods=['POST'])
@token_required
def update_user_settings(line_user_id):
    if not db:
        return jsonify({"status": "error", "message": "Database connection failed"}), 500

    data = request.get_json()
    if data is None:
        return jsonify({"status": "error", "message": "Invalid JSON"}), 400

    try:
        users_ref = db.collection('users')
        query = users_ref.where(filter=FieldFilter('line_user_id', '==', line_user_id)).limit(1)
        docs = list(query.stream())

        if not docs:
            return jsonify({"status": "error", "message": "User not found"}), 404

        user_doc_ref = docs[0].reference

        # Prepare settings update
        settings_update = {}
        if 'notifications_enabled' in data:
            settings_update['settings.notifications_enabled'] = bool(data['notifications_enabled'])

        if not settings_update:
            return jsonify({"status": "error", "message": "No valid settings provided"}), 400

        settings_update['updated_at'] = datetime.now().isoformat()
        user_doc_ref.update(settings_update)

        return jsonify({"status": "success", "message": "Settings updated successfully"}), 200

    except Exception as e:
        logger.error(f"Error updating settings for user {line_user_id}: {e}")
        return jsonify({"status": "error", "message": "Failed to update settings"}), 500

@app.route('/api/register_teacher', methods=['POST'])
@token_required
def register_teacher(line_user_id):
    """先生登録コードを検証し、ユーザーを先生として登録する"""
    if not db:
        return jsonify({"status": "error", "message": "Database connection failed"}), 500

    data = request.get_json()
    name = data.get('name')
    teacher_code = data.get('teacher_code')

    if not name or not teacher_code:
        return jsonify({"status": "error", "message": "名前と先生登録コードは必須です。"}), 400

    # 環境変数からマスター登録コードを取得
    master_code = os.getenv("TEACHER_REGISTRATION_CODE")
    if not master_code:
        logger.error("環境変数 'TEACHER_REGISTRATION_CODE' が設定されていません。")
        return jsonify({"status": "error", "message": "サーバー側で登録コードが設定されていません。"}), 500

    if teacher_code != master_code:
        return jsonify({"status": "error", "message": "先生登録コードが無効です。"}), 400

    try:
        users_ref = db.collection('users')
        query = users_ref.where(filter=FieldFilter('line_user_id', '==', line_user_id)).limit(1)
        docs = list(query.stream())

        if docs:
            # 既存ユーザーのロールを更新
            user_doc_ref = docs[0].reference
            user_doc_ref.update({
                'role': 'teacher',
                'name': name,
                'updated_at': datetime.now().isoformat()
            })
            logger.info(f"User {line_user_id} role updated to 'teacher'.")
        else:
            # 新規ユーザーを先生として作成
            new_user_data = {
                'line_user_id': line_user_id,
                'name': name,
                'role': 'teacher',
                'school': '',
                'class_name': '',
                'icon_path': '',
                'is_registered': True,
                'is_posting_diary': False,
                'created_at': datetime.now().isoformat(),
                'updated_at': datetime.now().isoformat()
            }
            users_ref.add(new_user_data)
            logger.info(f"New user {line_user_id} created with role 'teacher'.")

        return jsonify({"status": "success", "message": "先生として正常に登録されました。"}), 200

    except Exception as e:
        logger.error(f"Error during teacher registration for user {line_user_id}: {e}")
        return jsonify({"status": "error", "message": "登録処理中にエラーが発生しました。"}), 500

@app.route('/api/user/delete', methods=['DELETE'])
@token_required
def delete_user_account(line_user_id):
    if not db:
        return jsonify({"status": "error", "message": "Database connection failed"}), 500

    try:
        # This is a destructive operation. Proceed with caution.
        # Use a transaction to ensure atomicity if multiple dependent operations are needed.

        # 1. Find the user document
        users_ref = db.collection('users')
        user_query = users_ref.where(filter=FieldFilter('line_user_id', '==', line_user_id)).limit(1)
        user_docs = list(user_query.stream())

        if not user_docs:
            return jsonify({"status": "error", "message": "User not found"}), 404

        user_doc_ref = user_docs[0].reference

        # Batch delete related data
        batch = db.batch()

        # 2. Delete user's diaries
        diaries_query = db.collection('diaries').where(filter=FieldFilter('user_id', '==', line_user_id))
        for doc in diaries_query.stream():
            batch.delete(doc.reference)

        # 3. Delete user's comments
        comments_query = db.collection('comments').where(filter=FieldFilter('user_id', '==', line_user_id))
        for doc in comments_query.stream():
            batch.delete(doc.reference)

        # 4. Delete user's likes
        likes_query = db.collection('likes').where(filter=FieldFilter('user_id', '==', line_user_id))
        for doc in likes_query.stream():
            batch.delete(doc.reference)

        # 5. Delete the user document itself
        batch.delete(user_doc_ref)

        # Commit the batch
        batch.commit()

        # 6. Delete user's icon from Firebase Storage
        if bucket:
            blobs = bucket.list_blobs(prefix=f"{line_user_id}/")
            for blob in blobs:
                blob.delete()

        logger.info(f"Successfully deleted account and all data for user {line_user_id}")
        return jsonify({"status": "success", "message": "Account deleted successfully"}), 200

    except Exception as e:
        logger.error(f"Error deleting account for user {line_user_id}: {e}")
        return jsonify({"status": "error", "message": "Failed to delete account"}), 500


@app.route('/api/diaries', methods=['GET'])
@token_required
def get_diaries(requesting_line_user_id):
    if not db:
        return jsonify({"status": "error", "message": "Database connection failed"}), 500

    try:
        target_class_id = request.args.get('class_id')

        requesting_user_doc = db.collection('users').where(filter=FieldFilter('line_user_id', '==', requesting_line_user_id)).limit(1).get()
        if not requesting_user_doc:
            return jsonify({"status": "error", "message": "User not found"}), 404

        requesting_user_data = requesting_user_doc[0].to_dict()
        requesting_user_role = requesting_user_data.get('role', 'student')
        user_memberships = requesting_user_data.get('class_memberships', [])

        class_to_fetch_id = None
        if target_class_id:
            is_teacher_of_class = False
            if requesting_user_role == 'teacher':
                class_ref = db.collection('classes').document(target_class_id)
                class_doc = class_ref.get()
                if class_doc.exists and class_doc.to_dict().get('teacher_line_user_id') == requesting_line_user_id:
                    is_teacher_of_class = True

            is_approved_member = any(m.get('class_id') == target_class_id and m.get('status') == 'approved' for m in user_memberships)

            if not is_teacher_of_class and not is_approved_member:
                return jsonify({"status": "error", "message": "Unauthorized to view this class's diaries"}), 403

            class_to_fetch_id = target_class_id
        else:
            approved_memberships = [m for m in user_memberships if m.get('status') == 'approved']
            if approved_memberships:
                class_to_fetch_id = approved_memberships[0].get('class_id')

        if not class_to_fetch_id:
            return jsonify({"status": "success", "data": []}), 200

        # 日記を取得 (order_byを削除)
        diaries_ref = db.collection('diaries').where(filter=FieldFilter('class_id', '==', class_to_fetch_id))
        diaries_docs = list(diaries_ref.stream())

        # Python側でソート
        diaries_docs.sort(key=lambda doc: doc.to_dict().get('created_at', ''), reverse=True)

        diary_ids = [doc.id for doc in diaries_docs]
        likes_map = {}
        user_cache = {}

        # いいねを一括取得
        if diary_ids:
            for i in range(0, len(diary_ids), 30):
                chunk_ids = diary_ids[i:i+30]
                likes_query = db.collection('likes').where(filter=FieldFilter('diary_id', 'in', chunk_ids))
                for like in likes_query.stream():
                    diary_id = like.to_dict()['diary_id']
                    likes_map[diary_id] = likes_map.get(diary_id, 0) + 1

        user_likes_docs = db.collection('likes').where(filter=FieldFilter('user_id', '==', requesting_line_user_id)).stream()
        user_liked_diary_ids = {doc.to_dict()['diary_id'] for doc in user_likes_docs}

        diary_list = []
        for diary_doc in diaries_docs:
            diary_data = diary_doc.to_dict()
            user_id = diary_data.get('user_id')

            if user_id not in user_cache:
                user_doc_list = db.collection('users').where(filter=FieldFilter('line_user_id', '==', user_id)).limit(1).get()
                user_cache[user_id] = user_doc_list[0].to_dict() if user_doc_list else {}

            author_data = user_cache.get(user_id, {})
            author_name = author_data.get('name', '匿名ユーザー')

            if requesting_user_role != 'teacher':
                author_name = f"生徒-{user_id[-4:]}"

            like_count = likes_map.get(diary_doc.id, 0)
            is_liked_by_user = diary_doc.id in user_liked_diary_ids

            diary_list.append({
                'id': diary_doc.id,
                'author': author_name,
                'content': diary_data.get('content', ''),
                'created_at': diary_data.get('created_at', ''),
                'like_count': like_count,
                'is_liked_by_user': is_liked_by_user
            })

        return jsonify({"status": "success", "data": diary_list}), 200

    except Exception as e:
        logger.error(f"Error fetching diaries: {e}", exc_info=True)
        return jsonify({"status": "error", "message": "Failed to fetch diaries"}), 500

@app.route('/api/user/diaries/export', methods=['GET'])
@token_required
def export_user_diaries(line_user_id):
    if not db:
        return jsonify({"status": "error", "message": "Database connection failed"}), 500

    try:
        diaries_ref = db.collection('diaries').where(filter=FieldFilter('user_id', '==', line_user_id)).order_by('created_at', direction=firestore.Query.ASCENDING)
        diaries = diaries_ref.stream()

        diary_list = []
        for diary in diaries:
            diary_data = diary.to_dict()
            diary_list.append({
                'id': diary.id,
                'content': diary_data.get('content', ''),
                'created_at': diary_data.get('created_at', '')
            })

        return jsonify({"status": "success", "data": diary_list}), 200

    except Exception as e:
        logger.error(f"Error exporting diaries for user {line_user_id}: {e}")
        return jsonify({"status": "error", "message": "Failed to export diaries"}), 500

@app.route('/api/admin/user_role', methods=['POST'])
@token_required
def update_user_role(requesting_line_user_id):
    if not db:
        print("Firestore is not initialized.", file=sys.stderr)
        return jsonify({"status": "error", "message": "Database connection failed"}), 500

    data = request.get_json()
    target_user_id = data.get('target_user_id')
    new_role = data.get('new_role')

    if not target_user_id or not new_role:
        return jsonify({"status": "error", "message": "Missing target_user_id, or new_role"}), 400

    requesting_user_doc = db.collection('users').where(filter=FieldFilter('line_user_id', '==', requesting_line_user_id)).limit(1).get()
    requesting_user_data = requesting_user_doc[0].to_dict() if requesting_user_doc else {}

    if requesting_user_data.get('role') != 'teacher':
        return jsonify({"status": "error", "message": "Unauthorized: Only teachers can change user roles"}), 403

    try:
        users_ref = db.collection('users')
        query = users_ref.where(filter=FieldFilter('line_user_id', '==', target_user_id)).limit(1)
        docs = query.stream()

        user_doc_id = None
        for doc in docs:
            user_doc_id = doc.id
            break

        if user_doc_id:
            if new_role not in ['student', 'teacher']:
                return jsonify({"status": "error", "message": "Invalid role specified. Must be 'student' or 'teacher'."}), 400

            db.collection('users').document(user_doc_id).update({'role': new_role, 'updated_at': datetime.now().isoformat()})
            return jsonify({"status": "success", "message": f"User {target_user_id} role updated to {new_role}"}), 200
        else:
            return jsonify({"status": "error", "message": f"Target user {target_user_id} not found"}), 404

    except Exception as e:
        print(f"Error updating user role in Firestore: {e}", file=sys.stderr)
        return jsonify({"status": "error", "message": "Failed to update user role"}), 500

@app.route('/api/teacher/my_students', methods=['GET'])
@token_required
def get_my_students(teacher_line_user_id):
    if not db:
        print("Firestore is not initialized.", file=sys.stderr)
        return jsonify({"status": "error", "message": "Database connection failed"}), 500

    teacher_user_doc = db.collection('users').where(filter=FieldFilter('line_user_id', '==', teacher_line_user_id)).limit(1).get()
    teacher_user_data = teacher_user_doc[0].to_dict() if teacher_user_doc else {}

    if teacher_user_data.get('role') != 'teacher':
        return jsonify({"status": "error", "message": "Unauthorized: Only teachers can view their students"}), 403

    try:
        class_tokens_docs = db.collection('class_tokens').where(filter=FieldFilter('teacher_line_user_id', '==', teacher_line_user_id)).stream()
        teacher_class_token_ids = [doc.to_dict()['token_id'] for doc in class_tokens_docs]

        if not teacher_class_token_ids:
            return jsonify({"status": "success", "data": [], "message": "No classes or students found for this teacher"}), 200

        all_students = []
        for i in range(0, len(teacher_class_token_ids), 10):
            batch_token_ids = teacher_class_token_ids[i:i+10]
            students_docs = db.collection('users').where(filter=FieldFilter('class_token_id', 'in', batch_token_ids)).stream()
            for student_doc in students_docs:
                student_data = student_doc.to_dict()
                all_students.append({
                    'line_user_id': student_data.get('line_user_id'),
                    'name': student_data.get('name', '未登録'),
                    'school': student_data.get('school', '未登録'),
                    'class_name': student_data.get('class_name', '未登録'),
                    'icon_path': student_data.get('icon_path', ''),
                    'is_registered': student_data.get('is_registered', False),
                    'role': student_data.get('role', 'student')
                })

        return jsonify({"status": "success", "data": all_students}), 200

    except Exception as e:
        print(f"Error fetching students for teacher {teacher_line_user_id}: {e}", file=sys.stderr)
        return jsonify({"status": "error", "message": "Failed to fetch students"}), 500

@app.route('/api/teacher/class/<class_id>', methods=['GET'])
@token_required
def get_class_details(requesting_line_user_id, class_id):
    if not db:
        return jsonify({"status": "error", "message": "Database connection failed"}), 500

    try:
        class_ref = db.collection('classes').document(class_id)
        class_doc = class_ref.get()
        if not class_doc.exists:
            return jsonify({"status": "error", "message": "Class not found"}), 404

        class_data = class_doc.to_dict()
        teacher_line_user_id = class_data.get('teacher_line_user_id')

        if requesting_line_user_id != teacher_line_user_id:
            return jsonify({"status": "error", "message": "Unauthorized"}), 403

        teacher_user_doc = db.collection('users').where(filter=FieldFilter('line_user_id', '==', teacher_line_user_id)).limit(1).get()
        teacher_data = teacher_user_doc[0].to_dict() if teacher_user_doc else {}

        # 全ユーザーをストリーミングし、承認済みの生徒をフィルタリング
        all_users_stream = db.collection('users').stream()
        student_list = []
        for user_doc in all_users_stream:
            user_data = user_doc.to_dict()
            if class_id in user_data.get('approved_class_ids', []):
                student_list.append({
                    'name': user_data.get('name', '未登録'),
                    'icon_path': user_data.get('icon_path', ''),
                    'line_user_id': user_data.get('line_user_id')
                })

        response_data = {
            'class_name': class_data.get('class_name'),
            'teacher': {
                'name': teacher_data.get('name', '未登録'),
                'icon_path': teacher_data.get('icon_path', '')
            },
            'students': student_list
        }

        return jsonify({"status": "success", "data": response_data}), 200

    except Exception as e:
        logger.error(f"Error fetching class details for class {class_id}: {e}", exc_info=True)
        return jsonify({"status": "error", "message": "Failed to fetch class details"}), 500

@app.route('/api/teacher/class/<class_id>/students', methods=['GET'])
@token_required
def get_class_students(requesting_line_user_id, class_id):
    if not db:
        return jsonify({"status": "error", "message": "Database connection failed"}), 500

    try:
        class_ref = db.collection('classes').document(class_id)
        class_doc = class_ref.get()
        if not class_doc.exists:
            return jsonify({"status": "error", "message": "Class not found"}), 404

        class_data = class_doc.to_dict()
        teacher_line_user_id = class_data.get('teacher_line_user_id')

        if requesting_line_user_id != teacher_line_user_id:
            return jsonify({"status": "error", "message": "Unauthorized"}), 403

        # 全ユーザーをストリーミングし、承認済みの生徒をフィルタリング
        all_users_stream = db.collection('users').stream()
        student_list = []
        for user_doc in all_users_stream:
            user_data = user_doc.to_dict()
            if class_id in user_data.get('approved_class_ids', []):
                student_list.append({
                    'line_user_id': user_data.get('line_user_id'),
                    'name': user_data.get('name', '未登録'),
                    'school': user_data.get('school', '未登録'),
                    'icon_path': user_data.get('icon_path', '')
                })

        return jsonify({"status": "success", "data": student_list}), 200

    except Exception as e:
        logger.error(f"Error fetching students for class {class_id}: {e}", exc_info=True)
        return jsonify({"status": "error", "message": "Failed to fetch students"}), 500

@app.route('/api/teacher/pending_requests', methods=['GET'])
@token_required
def get_pending_requests(teacher_line_user_id):
    """指定されたクラスの承認待ち生徒一覧を取得する"""
    if not db:
        return jsonify({"status": "error", "message": "Database connection failed"}), 500

    class_id = request.args.get('class_id')
    if not class_id:
        return jsonify({"status": "error", "message": "class_id is required"}), 400

    try:
        # クラスの存在と先生の権限を確認
        class_ref = db.collection('classes').document(class_id)
        class_doc = class_ref.get()
        if not class_doc.exists or class_doc.to_dict().get('teacher_line_user_id') != teacher_line_user_id:
            return jsonify({"status": "error", "message": "Unauthorized or class not found"}), 403

        # 全ユーザーをストリーミングし、承認待ちの生徒をフィルタリング
        all_users_stream = db.collection('users').stream()
        pending_list = []
        for user_doc in all_users_stream:
            user_data = user_doc.to_dict()
            if class_id in user_data.get('pending_class_ids', []):
                pending_list.append({
                    'line_user_id': user_data.get('line_user_id'),
                    'name': user_data.get('name', '不明なユーザー'),
                    'icon_path': user_data.get('icon_path', ''),
                    'requested_at': next((m.get('requested_at') for m in user_data.get('class_memberships', []) if m.get('class_id') == class_id), None)
                })

        return jsonify({"status": "success", "data": pending_list}), 200

    except Exception as e:
        logger.error(f"Error fetching pending requests for class {class_id}: {e}", exc_info=True)
        return jsonify({"status": "error", "message": "Failed to fetch pending requests"}), 500

@app.route('/api/teacher/approve_request', methods=['POST'])
@token_required
def approve_request(teacher_line_user_id):
    """生徒のクラス参加を承認する"""
    if not db:
        return jsonify({"status": "error", "message": "Database connection failed"}), 500

    data = request.get_json()
    class_id = data.get('class_id')
    student_line_user_id = data.get('student_line_user_id')

    if not class_id or not student_line_user_id:
        return jsonify({"status": "error", "message": "class_id and student_line_user_id are required"}), 400

    try:
        # クラスの存在と先生の権限を確認
        class_ref = db.collection('classes').document(class_id)
        class_doc = class_ref.get()
        if not class_doc.exists or class_doc.to_dict().get('teacher_line_user_id') != teacher_line_user_id:
            return jsonify({"status": "error", "message": "Unauthorized or class not found"}), 403

        class_name = class_doc.to_dict().get('class_name', '不明なクラス')

        # 生徒のドキュメントを取得
        users_ref = db.collection('users')
        student_query = users_ref.where(filter=FieldFilter('line_user_id', '==', student_line_user_id)).limit(1)
        student_docs = list(student_query.stream())
        if not student_docs:
            return jsonify({"status": "error", "message": "Student not found"}), 404

        student_doc_ref = student_docs[0].reference
        student_data = student_docs[0].to_dict()

        # class_membershipsを更新
        memberships = student_data.get('class_memberships', [])
        updated_memberships = []
        found = False
        for m in memberships:
            if m.get('class_id') == class_id and m.get('status') == 'pending':
                m['status'] = 'approved'
                m['approved_at'] = datetime.now().isoformat()
                found = True
            updated_memberships.append(m)

        if not found:
            return jsonify({"status": "error", "message": "No pending request found for this student in this class"}), 404

        # Firestoreドキュメントを更新
        student_doc_ref.update({
            'class_memberships': updated_memberships,
            'pending_class_ids': firestore.ArrayRemove([class_id]),
            'approved_class_ids': firestore.ArrayUnion([class_id])
        })

        # LINEで通知を送信
        try:
            line_bot_api.push_message(
                student_line_user_id,
                TextSendMessage(text=f"おめでとうございます！「{class_name}」への参加が承認されました。")
            )
        except Exception as e:
            logger.error(f"Failed to send approval notification to {student_line_user_id}: {e}")
            # 通知が失敗しても、承認処理自体は成功として返す

        logger.info(f"Student {student_line_user_id} approved for class {class_id} by teacher {teacher_line_user_id}")
        return jsonify({"status": "success", "message": "Student approved successfully"}), 200

    except Exception as e:
        logger.error(f"Error approving request: {e}", exc_info=True)
        return jsonify({"status": "error", "message": "Failed to approve request"}), 500

@app.route('/api/teacher/reject_request', methods=['POST'])
@token_required
def reject_request(teacher_line_user_id):
    """生徒のクラス参加を拒否する"""
    if not db:
        return jsonify({"status": "error", "message": "Database connection failed"}), 500

    data = request.get_json()
    class_id = data.get('class_id')
    student_line_user_id = data.get('student_line_user_id')

    if not class_id or not student_line_user_id:
        return jsonify({"status": "error", "message": "class_id and student_line_user_id are required"}), 400

    try:
        # クラスの存在と先生の権限を確認
        class_ref = db.collection('classes').document(class_id)
        class_doc = class_ref.get()
        if not class_doc.exists or class_doc.to_dict().get('teacher_line_user_id') != teacher_line_user_id:
            return jsonify({"status": "error", "message": "Unauthorized or class not found"}), 403

        class_name = class_doc.to_dict().get('class_name', '不明なクラス')

        # 生徒のドキュメントを取得
        users_ref = db.collection('users')
        student_query = users_ref.where(filter=FieldFilter('line_user_id', '==', student_line_user_id)).limit(1)
        student_docs = list(student_query.stream())
        if not student_docs:
            return jsonify({"status": "error", "message": "Student not found"}), 404

        student_doc_ref = student_docs[0].reference
        student_data = student_docs[0].to_dict()

        # class_membershipsから該当のpendingリクエストを削除
        memberships = student_data.get('class_memberships', [])
        # 拒否対象ではないメンバーシップだけを残す
        updated_memberships = [m for m in memberships if not (m.get('class_id') == class_id and m.get('status') == 'pending')]

        if len(memberships) == len(updated_memberships):
             return jsonify({"status": "error", "message": "No pending request found for this student in this class"}), 404

        # Firestoreドキュメントを更新
        student_doc_ref.update({
            'class_memberships': updated_memberships,
            'pending_class_ids': firestore.ArrayRemove([class_id])
        })

        # LINEで通知を送信
        try:
            line_bot_api.push_message(
                student_line_user_id,
                TextSendMessage(text=f"「{class_name}」への参加申請は、今回は見送られました。")
            )
        except Exception as e:
            logger.error(f"Failed to send rejection notification to {student_line_user_id}: {e}")
            # 通知が失敗しても、拒否処理自体は成功として返す

        logger.info(f"Student {student_line_user_id} rejected for class {class_id} by teacher {teacher_line_user_id}")
        return jsonify({"status": "success", "message": "Student rejected successfully"}), 200

    except Exception as e:
        logger.error(f"Error rejecting request: {e}", exc_info=True)
        return jsonify({"status": "error", "message": "Failed to reject request"}), 500

@app.route('/api/teacher/remove_student', methods=['POST'])
@token_required
def remove_student(teacher_line_user_id):
    """生徒をクラスから削除する"""
    if not db:
        return jsonify({"status": "error", "message": "Database connection failed"}), 500

    data = request.get_json()
    class_id = data.get('class_id')
    student_line_user_id = data.get('student_line_user_id')

    if not class_id or not student_line_user_id:
        return jsonify({"status": "error", "message": "class_id and student_line_user_id are required"}), 400

    try:
        # クラスの存在と先生の権限を確認
        class_ref = db.collection('classes').document(class_id)
        class_doc = class_ref.get()
        if not class_doc.exists or class_doc.to_dict().get('teacher_line_user_id') != teacher_line_user_id:
            return jsonify({"status": "error", "message": "Unauthorized or class not found"}), 403

        class_name = class_doc.to_dict().get('class_name', '不明なクラス')

        # 生徒のドキュメントを取得
        users_ref = db.collection('users')
        student_query = users_ref.where(filter=FieldFilter('line_user_id', '==', student_line_user_id)).limit(1)
        student_docs = list(student_query.stream())
        if not student_docs:
            return jsonify({"status": "error", "message": "Student not found"}), 404

        student_doc_ref = student_docs[0].reference
        student_data = student_docs[0].to_dict()

        # class_membershipsから該当のapprovedメンバーシップを削除
        memberships = student_data.get('class_memberships', [])
        updated_memberships = [m for m in memberships if not (m.get('class_id') == class_id and m.get('status') == 'approved')]

        if len(memberships) == len(updated_memberships):
             return jsonify({"status": "error", "message": "No approved membership found for this student in this class"}), 404

        # Firestoreドキュメントを更新
        student_doc_ref.update({
            'class_memberships': updated_memberships,
            'approved_class_ids': firestore.ArrayRemove([class_id])
        })

        # LINEで通知を送信
        try:
            line_bot_api.push_message(
                student_line_user_id,
                TextSendMessage(text=f"「{class_name}」から退会させられました。")
            )
        except Exception as e:
            logger.error(f"Failed to send removal notification to {student_line_user_id}: {e}")
            # 通知が失敗しても、削除処理自体は成功として返す

        logger.info(f"Student {student_line_user_id} removed from class {class_id} by teacher {teacher_line_user_id}")
        return jsonify({"status": "success", "message": "Student removed successfully"}), 200

    except Exception as e:
        logger.error(f"Error removing student: {e}", exc_info=True)
        return jsonify({"status": "error", "message": "Failed to remove student"}), 500

# ==============================================================================
# Assignment Feature API Endpoints
# ==============================================================================

@app.route('/api/teacher/class/<class_id>/assignments', methods=['POST'])
@token_required
def create_assignment(teacher_line_user_id, class_id):
    """新しい課題を作成する"""
    if not db:
        return jsonify({"status": "error", "message": "Database connection failed"}), 500

    # 権限チェック：リクエスト者がそのクラスの先生であることを確認
    try:
        class_ref = db.collection('classes').document(class_id)
        class_doc = class_ref.get()
        if not class_doc.exists or class_doc.to_dict().get('teacher_line_user_id') != teacher_line_user_id:
            return jsonify({"status": "error", "message": "Unauthorized"}), 403
    except Exception as e:
        logger.error(f"Error checking class permissions: {e}")
        return jsonify({"status": "error", "message": "Failed to verify permissions"}), 500

    data = request.get_json()
    title = data.get('title')
    description = data.get('description', '')
    due_date = data.get('due_date') # ISO 8601 format string

    if not title or not due_date:
        return jsonify({"status": "error", "message": "Title and due date are required"}), 400

    try:
        new_assignment_ref = db.collection('assignments').document()
        new_assignment_data = {
            'id': new_assignment_ref.id,
            'class_id': class_id,
            'title': title,
            'description': description,
            'due_date': due_date,
            'created_at': datetime.now().isoformat()
        }
        new_assignment_ref.set(new_assignment_data)

        # Notify students in the class
        try:
            student_docs = db.collection('users').where(filter=FieldFilter('approved_class_ids', 'array_contains', class_id)).stream()
            student_line_ids = [doc.to_dict()['line_user_id'] for doc in student_docs]
            
            if student_line_ids:
                class_name = class_doc.to_dict().get('class_name', '')
                due_date_formatted = datetime.fromisoformat(due_date).strftime('%m月%d日 %H:%M')
                notification_text = f"【新しい課題のお知らせ】\nクラス「{class_name}」に新しい課題が追加されました。\n\n■ {title}\n期限: {due_date_formatted}\n\n「課題一覧」と送って確認してください。"
                line_bot_api.multicast(student_line_ids, TextSendMessage(text=notification_text))
                logger.info(f"Sent assignment notification to {len(student_line_ids)} students in class {class_id}.")

        except Exception as e:
            logger.error(f"Failed to send notification for new assignment {new_assignment_ref.id}: {e}")
            # Continue even if notification fails

        return jsonify({"status": "success", "message": "Assignment created successfully", "data": new_assignment_data}), 201

    except Exception as e:
        logger.error(f"Error creating assignment for class {class_id}: {e}")
        return jsonify({"status": "error", "message": "Failed to create assignment"}), 500

@app.route('/api/teacher/class/<class_id>/assignments', methods=['GET'])
@token_required
def get_assignments_for_class(teacher_line_user_id, class_id):
    """特定のクラスのすべての課題を取得する"""
    if not db:
        return jsonify({"status": "error", "message": "Database connection failed"}), 500

    # 権限チェック
    try:
        class_ref = db.collection('classes').document(class_id)
        class_doc = class_ref.get()
        if not class_doc.exists or class_doc.to_dict().get('teacher_line_user_id') != teacher_line_user_id:
            return jsonify({"status": "error", "message": "Unauthorized"}), 403
    except Exception as e:
        logger.error(f"Error checking class permissions: {e}")
        return jsonify({"status": "error", "message": "Failed to verify permissions"}), 500

    try:
        assignments_ref = db.collection('assignments').where(filter=FieldFilter('class_id', '==', class_id)).order_by('due_date', direction=firestore.Query.ASCENDING)
        docs = assignments_ref.stream()
        assignment_list = [doc.to_dict() for doc in docs]
        return jsonify({"status": "success", "data": assignment_list}), 200
    except Exception as e:
        logger.error(f"Error fetching assignments for class {class_id}: {e}")
        return jsonify({"status": "error", "message": "Failed to fetch assignments"}), 500

@app.route('/api/teacher/assignment/<assignment_id>', methods=['GET'])
@token_required
def get_assignment_detail(teacher_line_user_id, assignment_id):
    """特定の課題の詳細を取得する"""
    if not db:
        return jsonify({"status": "error", "message": "Database connection failed"}), 500

    try:
        assignment_ref = db.collection('assignments').document(assignment_id)
        assignment_doc = assignment_ref.get()
        if not assignment_doc.exists:
            return jsonify({"status": "error", "message": "Assignment not found"}), 404

        assignment_data = assignment_doc.to_dict()
        class_id = assignment_data.get('class_id')

        # 権限チェック
        class_ref = db.collection('classes').document(class_id)
        class_doc = class_ref.get()
        if not class_doc.exists or class_doc.to_dict().get('teacher_line_user_id') != teacher_line_user_id:
            return jsonify({"status": "error", "message": "Unauthorized"}), 403

        return jsonify({"status": "success", "data": assignment_data}), 200
    except Exception as e:
        logger.error(f"Error fetching assignment detail for {assignment_id}: {e}")
        return jsonify({"status": "error", "message": "Failed to fetch assignment details"}), 500


@app.route('/api/teacher/assignment/<assignment_id>/submissions', methods=['GET'])
@token_required
def get_submissions_for_assignment(teacher_line_user_id, assignment_id):
    """特定の課題のすべての提出物を取得する"""
    if not db:
        return jsonify({"status": "error", "message": "Database connection failed"}), 500

    try:
        # 権限チェックのために課題情報を取得
        assignment_ref = db.collection('assignments').document(assignment_id)
        assignment_doc = assignment_ref.get()
        if not assignment_doc.exists:
            return jsonify({"status": "error", "message": "Assignment not found"}), 404

        assignment_data = assignment_doc.to_dict()
        class_id = assignment_data.get('class_id')

        class_ref = db.collection('classes').document(class_id)
        class_doc = class_ref.get()
        if not class_doc.exists or class_doc.to_dict().get('teacher_line_user_id') != teacher_line_user_id:
            return jsonify({"status": "error", "message": "Unauthorized"}), 403

        # 提出物を取得
        submissions_ref = db.collection('submissions').where(filter=FieldFilter('assignment_id', '==', assignment_id)).order_by('submitted_at', direction=firestore.Query.DESCENDING)
        docs = submissions_ref.stream()
        
        submission_list = []
        user_cache = {}
        for doc in docs:
            submission_data = doc.to_dict()
            student_id = submission_data.get('student_line_user_id')

            # ユーザー情報をキャッシュから取得またはFirestoreから取得
            if student_id not in user_cache:
                user_query = db.collection('users').where(filter=FieldFilter('line_user_id', '==', student_id)).limit(1)
                user_docs = list(user_query.stream())
                if user_docs:
                    user_cache[student_id] = user_docs[0].to_dict()
                else:
                    user_cache[student_id] = {'name': '不明な生徒'} # 見つからない場合

            submission_data['student_name'] = user_cache[student_id].get('name', '不明な生徒')
            submission_list.append(submission_data)

        return jsonify({"status": "success", "data": submission_list}), 200
    except Exception as e:
        logger.error(f"Error fetching submissions for assignment {assignment_id}: {e}")
        return jsonify({"status": "error", "message": "Failed to fetch submissions"}), 500


@app.route('/api/teacher/class_analysis', methods=['GET'])
@token_required
def class_analysis(teacher_line_user_id):
    """クラス全体の活動を分析する"""
    if not db:
        return jsonify({"status": "error", "message": "Database connection failed"}), 500
    if not GEMINI_API_KEY:
        return jsonify({"status": "error", "message": "Gemini API key is not configured on the server."}), 500

    class_id = request.args.get('class_id')
    if not class_id:
        return jsonify({"status": "error", "message": "class_id is required"}), 400

    try:
        # 1. クラスの存在と先生の権限を確認
        class_ref = db.collection('classes').document(class_id)
        class_doc = class_ref.get()
        if not class_doc.exists or class_doc.to_dict().get('teacher_line_user_id') != teacher_line_user_id:
            return jsonify({"status": "error", "message": "Unauthorized or class not found"}), 403

        class_name = class_doc.to_dict().get('class_name', '不明なクラス')

        # 2. クラスに所属する承認済み生徒のLINE User IDを取得
        # approved_class_ids配列にclass_idが含まれるユーザーを検索
        student_docs = db.collection('users').where(filter=FieldFilter('approved_class_ids', 'array_contains', class_id)).stream()
        student_line_ids = [doc.to_dict()['line_user_id'] for doc in student_docs]

        if not student_line_ids:
            return jsonify({"status": "success", "analysis": "このクラスにはまだ承認済みの生徒がいません。"}), 200

        # 3. 生徒全員の日記データを取得
        all_diaries_content = ""
        # Firestoreの'in'クエリは最大10個の要素しか受け付けないため、分割して処理
        for i in range(0, len(student_line_ids), 10):
            batch_ids = student_line_ids[i:i+10]
            diaries_query = db.collection('diaries').where(filter=FieldFilter('user_id', 'in', batch_ids)).order_by('created_at').stream()
            for doc in diaries_query:
                data = doc.to_dict()
                all_diaries_content += f"ユーザーID: {data.get('user_id')}\n日付: {data.get('created_at')}\n内容: {data.get('content')}\n\n"

        if not all_diaries_content:
            return jsonify({"status": "success", "analysis": "このクラスの生徒はまだ日記を投稿していません。"}), 200

        # 4. Gemini APIでクラス全体の活動を分析
        prompt = f"""以下の日記群は、クラス「{class_name}」に所属する複数の生徒が書いたものです。
        これらの日記全体を分析し、クラス全体の活動や感情の傾向、共通のテーマなどを300字程度で要約してください。
        また、クラス全体でよく使われているキーワードを5つ抽出してください。

        # 出力形式
        ## クラス全体の要約
        ここに要約を記述

        ## 頻出キーワード
        - キーワード1
        - キーワード2
        ...

        # 日記データ
        {all_diaries_content}
        """

        model = genai.GenerativeModel('gemini-pro')
        response = model.generate_content(prompt)

        return jsonify({"status": "success", "analysis": response.text}), 200

    except Exception as e:
        logger.error(f"Error during class analysis: {e}", exc_info=True)
        return jsonify({"status": "error", "message": f"Failed to perform class analysis: {str(e)}"}), 500

@app.route('/api/teacher/student_report', methods=['GET'])
@token_required
def student_report(teacher_line_user_id):
    """個別生徒の活動レポートを生成する"""
    if not db:
        return jsonify({"status": "error", "message": "Database connection failed"}), 500

    class_id = request.args.get('class_id')
    student_line_user_id = request.args.get('student_line_user_id')

    if not class_id or not student_line_user_id:
        return jsonify({"status": "error", "message": "class_id and student_line_user_id are required"}), 400

    try:
        # 1. クラスの存在と先生の権限を確認
        class_ref = db.collection('classes').document(class_id)
        class_doc = class_ref.get()
        if not class_doc.exists or class_doc.to_dict().get('teacher_line_user_id') != teacher_line_user_id:
            return jsonify({"status": "error", "message": "Unauthorized or class not found"}), 403

        class_name = class_doc.to_dict().get('class_name', '不明なクラス')

        # 2. 生徒がそのクラスに所属しているか確認
        student_query = db.collection('users').where(filter=FieldFilter('line_user_id', '==', student_line_user_id)).limit(1)
        student_docs = list(student_query.stream())
        if not student_docs:
            return jsonify({"status": "error", "message": "Student not found"}), 404

        student_data = student_docs[0].to_dict()
        if class_id not in student_data.get('approved_class_ids', []):
            return jsonify({"status": "error", "message": "Student is not approved for this class"}), 403

        student_name = student_data.get('display_name', '不明な生徒')

        # 3. 生徒の日記データを取得
        diaries_query = db.collection('diaries').where(filter=FieldFilter('user_id', '==', student_line_user_id)).order_by('created_at').stream()
        diaries = []
        total_word_count = 0
        latest_post_date = None

        for doc in diaries_query:
            diary_data = doc.to_dict()
            diaries.append(diary_data)
            if 'content' in diary_data:
                total_word_count += len(diary_data['content'])
            if 'created_at' in diary_data:
                if latest_post_date is None or diary_data['created_at'] > latest_post_date:
                    latest_post_date = diary_data['created_at']

        total_posts = len(diaries)
        average_word_count = total_word_count / total_posts if total_posts > 0 else 0

        report_data = {
            "student_name": student_name,
            "class_name": class_name,
            "total_posts": total_posts,
            "total_word_count": total_word_count,
            "average_word_count": round(average_word_count),
            "latest_post_date": latest_post_date,
            "diaries_content": "\n\n".join([f"日付: {d.get('created_at')}\n内容: {d.get('content')}" for d in diaries])
        }

        # Gemini APIで日記全体の要約を生成 (オプション)
        gemini_summary = "日記がありません。"
        if report_data["diaries_content"]:
            prompt = f"""以下の日記は、生徒「{student_name}」がクラス「{class_name}」で書いたものです。
            これらの日記全体を分析し、生徒の活動や感情の傾向、学習への取り組みなどを200字程度で要約してください。
            先生が指導に役立てられるような、客観的かつ建設的な視点で記述してください。

            # 日記データ
            {report_data["diaries_content"]}
            """
            model = genai.GenerativeModel('gemini-pro')
            response = model.generate_content(prompt)
            gemini_summary = response.text

        report_data["gemini_summary"] = gemini_summary

        return jsonify({"status": "success", "report": report_data}), 200

    except Exception as e:
        logger.error(f"Error generating student report: {e}", exc_info=True)
        return jsonify({"status": "error", "message": f"Failed to generate student report: {str(e)}"}), 500

@app.route('/api/home_analysis_summary', methods=['GET'])
@token_required
def home_analysis_summary(line_user_id):
    if not GEMINI_API_KEY:
        return jsonify({"status": "error", "message": "Gemini API key is not configured on the server."}), 500

    try:
        # 1. Fetch user's diaries
        diaries_query = db.collection('diaries').where(filter=FieldFilter('user_id', '==', line_user_id))
        all_diaries_docs = list(diaries_query.stream())

        # Filter by date in Python
        start_date = datetime.now() - timedelta(days=30)
        recent_diaries_docs = [
            doc for doc in all_diaries_docs
            if doc.to_dict().get('created_at') and doc.to_dict().get('created_at') >= start_date.isoformat()
        ]

        recent_diaries_docs.sort(key=lambda doc: doc.to_dict().get('created_at', ''))

        if len(recent_diaries_docs) < 3: # 日記が少なすぎる場合は分析しない
             return jsonify({"summary": "直近30日の日記が3件未満のため、分析は行いませんでした。日記を3件以上投稿すると、活動のサマリーが表示されます。"}), 200

        diaries_content = ""
        for doc in recent_diaries_docs:
            data = doc.to_dict()
            diaries_content += f"日付: {data.get('created_at')}\n内容: {data.get('content')}\n\n"

        # 2. Create a concise prompt
        prompt = f"""以下の日記群から、筆者の活動や感情の傾向を200字程度で簡潔に要約してください。ポジティブな側面に焦点を当て、本人を励ますようなトーンで記述してください。

# 日記データ
{diaries_content}
"""
        # 3. Call Gemini API
        model = genai.GenerativeModel('gemini-pro')
        response = model.generate_content(prompt)

        return jsonify({"summary": response.text}), 200

    except Exception as e:
        logger.error(f"Error during home analysis summary: {e}", exc_info=True)
        return jsonify({"status": "error", "message": f"Failed to perform analysis: {str(e)}"}), 500

@app.route('/api/gemini_analysis', methods=['GET'])
@token_required
def gemini_analysis(line_user_id):
    if not GEMINI_API_KEY:
        return jsonify({"status": "error", "message": "Gemini API key is not configured on the server."}), 500

    try:
        analysis_type = request.args.get('type')
        period_months = request.args.get('period')

        # 1. Fetch all user's diaries first
        diaries_query = db.collection('diaries').where(filter=FieldFilter('user_id', '==', line_user_id))
        all_diaries_docs = list(diaries_query.stream())

        # 2. Filter by period in Python
        target_diaries_docs = []
        if period_months and period_months != 'all':
            try:
                months = int(period_months)
                start_date = datetime.now() - timedelta(days=months * 30)
                target_diaries_docs = [
                    doc for doc in all_diaries_docs
                    if doc.to_dict().get('created_at') and doc.to_dict().get('created_at') >= start_date.isoformat()
                ]
            except ValueError:
                return jsonify({"status": "error", "message": "Invalid period format."}), 400
        else:
            target_diaries_docs = all_diaries_docs

        target_diaries_docs.sort(key=lambda doc: doc.to_dict().get('created_at', ''))

        diaries_content = ""
        for doc in target_diaries_docs:
            data = doc.to_dict()
            diaries_content += f"日付: {data.get('created_at')}\n内容: {data.get('content')}\n\n"

        if not diaries_content:
            return jsonify({"analysis": "分析対象の日記がありません。"}), 200

        # 2. Select prompt based on analysis type
        prompt = ""
        if analysis_type == 'summary':
            prompt = f"""以下の日記群全体を要約し、筆者の活動や感情の傾向を300字程度でまとめてください。また、その内容を代表する重要なキーワードを10個抽出してください。

# 出力形式
## 要約
ここに要約を記述

## キーワード
- キーワード1
- キーワード2
...

# 日記データ
{diaries_content}
"""
        elif analysis_type == 'star':
            prompt = f"""以下の日記群から、筆者の「強み」や「改善点」が推測できるエピソードを1つ選び出し、STARメソッド（Situation, Task, Action, Result）の形式で整理してください。そして、そのエピソードから導き出される客観的な強みと、今後の改善点を具体的に分析してください。

# 出力形式
## エピソード分析 (STARメソッド)
- **Situation（状況）:** どのような状況でしたか。
- **Task（課題）:** どのような課題や目標がありましたか。
- **Action（行動）:** それに対して、あなたは具体的にどう行動しましたか。
- **Result（結果）:** 行動の結果、どうなりましたか。

## 分析結果
- **あなたの強み:** 上記のエピソードから、〇〇という強みが推測されます。なぜなら...
- **今後の改善点:** より成長するために、〇〇という点を意識すると良いでしょう。なぜなら...

# 日記データ
{diaries_content}
"""
        elif analysis_type == 'values':
            prompt = f"""以下の日記群から、筆者が仕事や人生において大切にしている「価値観」や、特に「興味を持っている分野」を3つ推定してください。それぞれの価値観・興味について、その根拠となる日記の記述を引用し、なぜそう判断したのかを簡潔に説明してください。

# 出力形式
## 価値観・興味の分析
### 1. 価値観/興味1 (例: チームでの目標達成)
- **根拠:** (日記の具体的な記述を1,2文引用)
- **分析:** この記述から、一人で完結する作業よりも、他者と協力して何かを成し遂げることに喜びを感じる傾向が読み取れます。

### 2. 価値観/興味2 (例: 新しい技術への探求心)
- **根拠:** (日記の具体的な記述を1,2文引用)
- **分析:** ...

### 3. 価値観/興味3 (...)
- **根拠:** ...
- **分析:** ...

# 日記データ
{diaries_content}
"""
        else:
            return jsonify({"status": "error", "message": "Invalid analysis type."}), 400

        # 3. Call Gemini API
        model = genai.GenerativeModel('gemini-pro')
        response = model.generate_content(prompt)

        return jsonify({"analysis": response.text}), 200

    except Exception as e:
        logger.error(f"Error during Gemini analysis: {e}", exc_info=True)
        return jsonify({"status": "error", "message": f"Failed to perform analysis: {str(e)}"}), 500

@app.route('/api/diaries/<diary_id>/like', methods=['POST'])
@token_required
def like_diary(liking_user_id, diary_id):
    if not db:
        print("Firestore is not initialized.", file=sys.stderr)
        return jsonify({"status": "error", "message": "Database connection failed"}), 500

    try:
        diary_ref = db.collection('diaries').document(diary_id)
        if not diary_ref.get().exists:
            return jsonify({"status": "error", "message": "Diary not found"}), 404

        likes_ref = db.collection('likes')
        existing_like_query = likes_ref.where(filter=FieldFilter('diary_id', '==', diary_id)).where(filter=FieldFilter('user_id', '==', liking_user_id)).limit(1)
        existing_like = existing_like_query.get()

        if existing_like:
            for doc in existing_like:
                likes_ref.document(doc.id).delete()
            return jsonify({"status": "success", "message": "Like removed"}), 200
        else:
            likes_ref.add({
                'diary_id': diary_id,
                'user_id': liking_user_id,
                'created_at': datetime.now().isoformat()
            })
            return jsonify({"status": "success", "message": "Like added"}), 201

    except Exception as e:
        print(f"Error processing like for diary {diary_id}: {e}", file=sys.stderr)
        return jsonify({"status": "error", "message": "Failed to process like"}), 500

@app.route('/api/diaries/<diary_id>/comment', methods=['POST'])
@token_required
def add_comment(commenting_user_id, diary_id):
    if not db:
        print("Firestore is not initialized.", file=sys.stderr)
        return jsonify({"status": "error", "message": "Database connection failed"}), 500

    data = request.get_json()
    comment_content = data.get('content')

    if not comment_content:
        return jsonify({"status": "error", "message": "Missing comment content"}), 400

    try:
        # OpenAI API for tag generation
        if openai.api_key:
            prompt = f"""次の文章からポジティブなタグのみを3つまで抽出してください（複数ある場合はカンマ区切り）:
{comment_content}"""
            response = openai.ChatCompletion.create(
                model="gpt-3.5-turbo",
                messages=[{"role": "user", "content": prompt}],
                max_tokens=60,
                temperature=0
            )
            tags_text = response.choices[0].message.content.strip()
            tags = [tag.strip() for tag in tags_text.split(",") if tag.strip()]
        else:
            tags = []
            logger.warning("OpenAI API key not set. Skipping tag generation.")


        diary_ref = db.collection('diaries').document(diary_id)
        if not diary_ref.get().exists:
            return jsonify({"status": "error", "message": "Diary not found"}), 404

        for ng_word in NG_WORDS:
            if ng_word in comment_content:
                return jsonify({"status": "error", "message": f"""不適切な言葉が含まれています。コメントは投稿されませんでした。
「{ng_word}」のような言葉は使用しないでください。"""}), 400

        db.collection('comments').add({
            'diary_id': diary_id,
            'user_id': commenting_user_id,
            'content': comment_content,
            'tags': tags,
            'created_at': datetime.now().isoformat()
        })
        return jsonify({"status": "success", "message": "Comment added"}), 201

    except Exception as e:
        print(f"Error adding comment for diary {diary_id}: {e}", file=sys.stderr)
        return jsonify({"status": "error", "message": "Failed to add comment"}), 500

@app.route('/api/diary-tags', methods=['GET'])
@token_required
def diary_tags(line_user_id):
    if not db:
        return jsonify({"status": "error", "message": "Database connection failed"}), 500

    try:
        personal_counts = {}
        class_counts = {}

        # --- Personal Aggregation (Tags from comments on user's own diaries) ---
        user_diaries_ref = db.collection('diaries').where(filter=FieldFilter('user_id', '==', line_user_id))
        user_diary_docs = list(user_diaries_ref.stream())
        user_diary_ids = [doc.id for doc in user_diary_docs]

        if user_diary_ids:
            all_personal_comments = []
            for i in range(0, len(user_diary_ids), 30):
                chunk_ids = user_diary_ids[i:i+30]
                comments_query = db.collection('comments').where(filter=FieldFilter('diary_id', 'in', chunk_ids))
                all_personal_comments.extend(comments_query.stream())

            for comment in all_personal_comments:
                for tag in comment.to_dict().get('tags', []):
                    personal_counts[tag] = personal_counts.get(tag, 0) + 1

        # --- Class Aggregation ---
        class_id = request.args.get('class_id')
        if class_id:
            # ユーザーがそのクラスのメンバーであることを確認（セキュリティチェック）
            user_doc_query = db.collection('users').where(filter=FieldFilter('line_user_id', '==', line_user_id)).limit(1)
            user_doc = list(user_doc_query.get())
            is_member = False
            if user_doc:
                user_data = user_doc[0].to_dict()
                if any(m.get('class_id') == class_id and m.get('status') == 'approved' for m in user_data.get('class_memberships', [])):
                    is_member = True
                if user_data.get('role') == 'teacher':
                    class_doc_ref = db.collection('classes').document(class_id).get()
                    if class_doc_ref.exists and class_doc_ref.to_dict().get('teacher_line_user_id') == line_user_id:
                        is_member = True # 先生もメンバーとみなす

            if is_member:
                class_diaries_ref = db.collection('diaries').where(filter=FieldFilter('class_id', '==', class_id))
                class_diary_docs = list(class_diaries_ref.stream())
                class_diary_ids = [doc.id for doc in class_diary_docs]

                if class_diary_ids:
                    all_class_comments = []
                    for i in range(0, len(class_diary_ids), 30):
                        chunk_ids = class_diary_ids[i:i+30]
                        comments_query = db.collection('comments').where(filter=FieldFilter('diary_id', 'in', chunk_ids))
                        all_class_comments.extend(comments_query.stream())

                    for comment in all_class_comments:
                        for tag in comment.to_dict().get('tags', []):
                            class_counts[tag] = class_counts.get(tag, 0) + 1

        return jsonify({
            "status": "success",
            "data": {
                "personal": personal_counts,
                "class_agg": class_counts,
                "school_agg": {} # School aggregation is disabled
            }
        }), 200

    except Exception as e:
        logger.error(f"Error generating diary-tags: {e}", exc_info=True)
        return jsonify({"status": "error", "message": "Failed to generate tags"}), 500



@app.route('/api/diaries/<diary_id>/comments', methods=['GET'])
@token_required
def get_comments(requesting_line_user_id, diary_id):
    if not db:
        print("Firestore is not initialized.", file=sys.stderr)
        return jsonify({"status": "error", "message": "Database connection failed"}), 500

    requesting_user_doc_query = db.collection('users').where(filter=FieldFilter('line_user_id', '==', requesting_line_user_id)).limit(1)
    requesting_user_doc = list(requesting_user_doc_query.get())
    requesting_user_data = requesting_user_doc[0].to_dict() if requesting_user_doc else {}
    requesting_user_role = requesting_user_data.get('role', 'student')

    try:
        diary_ref = db.collection('diaries').document(diary_id)
        if not diary_ref.get().exists:
            return jsonify({"status": "error", "message": "Diary not found"}), 404

        comments_ref = db.collection('comments').where(filter=FieldFilter('diary_id', '==', diary_id)).order_by('created_at')
        comments = comments_ref.stream()

        comment_list = []
        user_cache = {} # ユーザー情報をキャッシュしてFirestoreへのアクセスを減らす

        for comment in comments:
            comment_data = comment.to_dict()
            user_id = comment_data.get('user_id')

            if user_id not in user_cache:
                user_doc_query = db.collection('users').where(filter=FieldFilter('line_user_id', '==', user_id)).limit(1)
                user_doc = list(user_doc_query.get())
                user_cache[user_id] = user_doc[0].to_dict() if user_doc else {}

            author_data = user_cache.get(user_id, {})

            # 先生の場合は実名、それ以外は匿名表示
            author_name = author_data.get('name', '匿名ユーザー')
            if requesting_user_role != 'teacher':
                author_name = f"生徒-{user_id[-4:]}"

            comment_list.append({
                'id': comment.id,
                'author': author_name,
                'content': comment_data.get('content', ''),
                'created_at': comment_data.get('created_at', '')
            })

        return jsonify({"status": "success", "data": comment_list}), 200

    except Exception as e:
        print(f"Error fetching comments for diary {diary_id}: {e}", file=sys.stderr)
        return jsonify({"status": "error", "message": "Failed to fetch comments"}), 500
from openpyxl import load_workbook

def fill_resume(data, template_path="A4_format.xlsx", output_path="output.xlsx"):
    """
    data(dict) に以下が含まれる想定：
    {
        "furigana": "...",
        "name": "...",
        "birthday": "...",
        "address_kana1": "...",
        "zip1": "...",
        "address1": "...",
        "tel1": "...",
        "email1": "...",
        "address_kana2": "...",
        "zip2": "...",
        "address2": "...",
        "tel2": "...",
        "email2": "...",
        "education": [ {"year": "", "month": "", "text": "" }, ... ],
        "licenses": [ {"year": "", "month": "", "text": "" }, ... ],
        "motivation": "...",
        "notes": "..."
    }
    """

    wb = load_workbook(template_path)
    ws = wb.active

    # ----------------------------
    # 基本情報
    # ----------------------------
    ws["C6"] = data.get("furigana", "")
    ws["C9"] = data.get("name", "")
    ws["B14"] = data.get("birthday", "")

    ws["C16"] = data.get("address_kana1", "")
    ws["C19"] = data.get("zip1", "")
    ws["C21"] = data.get("address1", "")
    ws["I16"] = data.get("tel1", "")
    ws["H21"] = data.get("email1", "")

    ws["C25"] = data.get("address_kana2", "")
    ws["C28"] = data.get("zip2", "")
    ws["C30"] = data.get("address2", "")
    ws["I25"] = data.get("tel2", "")
    ws["H30"] = data.get("email2", "")

    # ----------------------------
    # 学歴・職歴
    # ----------------------------
    year_cells = ["B38","B41","B44","B47","B50","B53","B56","B59","B62","B65","B68","B71","B74","B77","B80","B83",
                  "L5","L8","L11","L14","L16","L19"]
    month_cells = ["C38","C41","C44","C47","C50","C53","C56","C59","C62","C65","C68","C71","C74","C77","C80","C83",
                   "M5","M8","M11","M14","M16","M19"]
    text_cells = ["D38","D41","D44","D47","D50","D53","D56","D59","D62","D65","D68","D71","D74","D77","D80","D83",
                  "N5","N8","N11","N14","N16","N19"]

    education = data.get("education", [])
    for i, item in enumerate(education):
        if i >= len(year_cells): break
        ws[year_cells[i]] = item.get("year", "")
        ws[month_cells[i]] = item.get("month", "")
        ws[text_cells[i]] = item.get("text", "")

    # ----------------------------
    # 資格・免許
    # ----------------------------
    lic_year = ["L25","L28","L31","L34","L37","L40"]
    lic_month = ["M25","M28","M31","M34","M37","M40"]
    lic_text = ["N25","N28","N31","N34","N37","N40"]

    licenses = data.get("licenses", [])
    for i, item in enumerate(licenses):
        if i >= len(lic_year): break
        ws[lic_year[i]] = item.get("year", "")
        ws[lic_month[i]] = item.get("month", "")
        ws[lic_text[i]] = item.get("text", "")

    # ----------------------------
    # 志望動機（1セル）
    # ----------------------------
    ws["L47"] = data.get("motivation", "")

    # ----------------------------
    # 本人希望記入欄（複数行 OK）
    # 行ごとに割り当てる
    # ----------------------------
    notes_lines = data.get("notes", "").split("\n")
    notes_cells = ["L71","L74","L77","L80","L83"]

    for i, line in enumerate(notes_lines):
        if i >= len(notes_cells): break
        ws[notes_cells[i]] = line

    # ----------------------------
    # 保存
    # ----------------------------
    wb.save(output_path)

@app.route('/resume')
def resume_form():
    return render_template("resume_form.html")

@app.route('/resume/create', methods=['POST'])
def resume_create():
    try:
        payload = request.get_json(force=True)

        # 必要ならここで認可チェック（ログインしているユーザーだけ許可、など）
        # 例: if not current_user: abort(401)

        # fill_resume() は先に渡した関数を想定
        # テンプレ名は "A4_format.xlsx" と仮定（同ディレクトリ or フルパスで）
        tmp_out = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        tmp_out.close()
        output_path = tmp_out.name

        # fill_resume 関数を使って Excel を書き出す
        template_path = os.path.join(BASE_DIR, "A4_format.xlsx")
        fill_resume(payload, template_path=template_path, output_path=output_path)

        # 返却（ダウンロード）
        filename = f"{payload.get('name','resume')}_resume.xlsx"
        return send_file(output_path, as_attachment=True, download_name=filename)

    except Exception as e:
        app.logger.exception("resume_create error")
        return ("Server error: " + str(e)), 500
    finally:
        # send_file が終わったら temp を消すのは send_file 実行後に行う必要あり。
        # ここでは残しておき、定期クリーンなどで消すのを推奨。
        pass
# ==============================================================================
# Page Rendering
# ==============================================================================
@app.route('/')
def index():
    return render_template('index.html', liff_id=LIFF_ID_PRIMARY)

@app.route('/mypage')
def mypage():
    return render_template('mypage.html', liff_id=LIFF_ID_PRIMARY)

@app.route('/settings')
def settings():
    return render_template('settings.html', liff_id=LIFF_ID_PRIMARY)

@app.route('/rules')
def rules():
    return render_template('rules.html', liff_id=LIFF_ID_PRIMARY)

@app.route('/contact')
def contact():
    return render_template('contact.html', liff_id=LIFF_ID_PRIMARY)

@app.route('/register_teacher')
def register_teacher_page():
    return render_template('register_teacher.html', liff_id=LIFF_ID_PRIMARY)

@app.route('/teacher_dashboard')
def teacher_dashboard():
    return render_template('teacher_dashboard.html', liff_id=LIFF_ID_PRIMARY)

@app.route('/class_home')
def class_home():
    return render_template('class_home.html', liff_id=LIFF_ID_PRIMARY)

@app.route('/class_detail')
def class_detail():
    return render_template('class_detail.html', liff_id=LIFF_ID_PRIMARY)

@app.route('/join_class')
def join_class_page():
    return render_template('join_class.html', liff_id=LIFF_ID_PRIMARY)

@app.route('/posts')
def posts():
    return render_template('posts.html', liff_id=LIFF_ID_PRIMARY)

@app.route('/analysis')
def analysis():
    return render_template('analysis.html', liff_id=LIFF_ID_PRIMARY)

@app.route('/resume_form')
def resume_form():
    return render_template('resume_form.html', liff_id=LIFF_ID_PRIMARY)

@app.route('/resume_view')
def resume_view():
    return render_template('resume_view.html', liff_id=LIFF_ID_PRIMARY)

@app.route('/teacher/assignments')
@token_required
def teacher_assignments_page(line_user_id):
    # Role check can be done on the client-side based on /api/user or here
    return render_template('teacher_assignments.html', liff_id=LIFF_ID_PRIMARY)

@app.route('/teacher/assignment_detail')
@token_required
def assignment_detail_page(line_user_id):
    return render_template('assignment_detail.html', liff_id=LIFF_ID_PRIMARY)